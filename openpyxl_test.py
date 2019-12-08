from openpyxl import Workbook, load_workbook
#wb = Workbook()
#ws = wb.active
#c = ws['A4']
#c.value = 10
#print(ws.max_row)
#wb.save('test.xlsx')
#print('saved')

wb2 = load_workbook('test.xlsx')
ws = wb2.active