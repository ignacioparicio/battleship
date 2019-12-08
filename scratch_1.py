from openpyxl import Workbook, load_workbook
wb = Workbook()
ws = wb.active
c = ws['A4']
c.value = 10
wb.save('balances.xlsx')

#wb2 = load_workbook('test.xlsx')